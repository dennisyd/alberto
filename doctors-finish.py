import os
from openpyxl import load_workbook
import warnings
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import NamedStyle

# Suppress warnings
warnings.filterwarnings('ignore', category=UserWarning)

def process_excel_file(file_path):
    try:
        # Load workbook with minimal options
        wb = load_workbook(filename=file_path)
        
        # Check for 'Info' sheet (case insensitive)
        info_sheet_name = None
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == 'info':
                info_sheet_name = sheet_name
                break
        
        if info_sheet_name:
            ws = wb[info_sheet_name]
            
            # Store the sheet's hidden state
            was_hidden = ws.sheet_state != 'visible'
            
            # Unhide the sheet temporarily
            ws.sheet_state = 'visible'
            
            # Find first blank row in column A
            first_blank_row = None
            for row in range(1, ws.max_row + 2):
                if ws[f'A{row}'].value is None:
                    first_blank_row = row
                    break
            
            if first_blank_row:
                # Add 'VSAT' in column A
                ws[f'A{first_blank_row}'] = 'VSAT'
                # Add specified text in column B
                ws[f'B{first_blank_row}'] = '93923 x2/93040/95923/95924'
            
            # Handle column E
            col_e_values = []
            last_row = ws.max_row
            
            # Collect existing values in column E, starting from row 2 to skip the header
            for row in range(2, last_row + 1):
                val = ws[f'E{row}'].value
                if val:
                    col_e_values.append(val)
            
            # Add 'Finkelstein' if not already present
            if 'Finkelstein' not in col_e_values:
                col_e_values.append('Finkelstein')
            
            # Sort values
            col_e_values.sort()
            
            # Write back sorted values, starting from row 2
            for idx, value in enumerate(col_e_values, 2):
                ws[f'E{idx}'] = value
            
            # Restore the sheet's hidden state if it was hidden
            if was_hidden:
                ws.sheet_state = 'hidden'
        else:
            print(f"Warning: 'Info' sheet not found in {file_path}")
        
        # Process sheets containing '2024' and add validation to renamed sheets
        for sheet_name in wb.sheetnames:
            if '2024' in sheet_name:
                new_name = sheet_name.replace('2024', '2025')
                sheet = wb[sheet_name]
                print(f"Adding validations and formulas to sheet: {sheet_name} (will be renamed to {new_name})")
                
                # Create data validation for column F (VSAT validation)
                dv_f = DataValidation(
                    type="list",
                    formula1="=INFO!$A$2:$A$23",
                    allow_blank=True
                )
                dv_f.error = 'Your entry is not in the list'
                dv_f.errorTitle = 'Invalid Entry'
                dv_f.prompt = 'Please select a value from the list'
                dv_f.promptTitle = 'Select a Value'
                
                # Create data validation for column L (Reading Doctor validation)
                dv_l = DataValidation(
                    type="list",
                    formula1="=INFO!$E$2:$E$10",
                    allow_blank=True
                )
                dv_l.error = 'Your entry is not in the list'
                dv_l.errorTitle = 'Invalid Entry'
                dv_l.prompt = 'Please select a Reading Doctor'
                dv_l.promptTitle = 'Select Doctor'
                
                # Create data validation for column M (Tech validation)
                dv_m = DataValidation(
                    type="list",
                    formula1="=INFO!$D$2:$D$4",
                    allow_blank=True
                )
                dv_m.error = 'Your entry is not in the list'
                dv_m.errorTitle = 'Invalid Entry'
                dv_m.prompt = 'Please select a Tech'
                dv_m.promptTitle = 'Select Tech'
                
                try:
                    # Add validation ranges
                    dv_f.add('F9:F3020')
                    dv_l.add('L9:L3020')
                    dv_m.add('M9:M3020')
                    
                    # Add validations to the sheet
                    sheet.add_data_validation(dv_f)
                    sheet.add_data_validation(dv_l)
                    sheet.add_data_validation(dv_m)
                    
                    # Add formulas to column G
                    for row in range(9, 3021):  # G9:G3020
                        sheet[f'G{row}'] = f'=IFERROR(VLOOKUP(F{row},INFO!$A$2:$B$22,2,FALSE),"")'
                    
                    # Add formulas to S7, S8, and T8
                    sheet['S7'] = '=INFO!E10'
                    sheet['T7'] = '=COUNTIF($L$9:$L$3020,S7)'
                    sheet['U7'] = '=SUMIFS($K$9:$K$3020,$L$9:$L$3020,S7)'

		    # Format U7 as currency
                    currency_format = NamedStyle(name='currency_style', number_format='$#,##0.00')
                    sheet['U7'].style = currency_format
                    
                    print(f"All validations and formulas added to sheet successfully")
                    
                except Exception as validation_error:
                    print(f"Error adding validation or formulas: {str(validation_error)}")
                
                # Rename the sheet
                sheet.title = new_name
                print(f"Renamed sheet to: {new_name}")
        
        # Save the workbook
        print(f"Saving workbook: {file_path}")
        wb.save(file_path)
        print(f"Successfully processed: {file_path}")
        wb.close()
        return True
        
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        if 'wb' in locals():
            wb.close()
        return False

def main():
    # Directory path
    directory = r"C:\Users\denni\Downloads\Alberto 2024\2025"
    
    # Ensure directory exists
    if not os.path.exists(directory):
        print(f"Directory not found: {directory}")
        return
    
    # Process all Excel files in the directory
    success_count = 0
    failure_count = 0
    
    print(f"\nStarting Excel file processing in: {directory}\n")
    
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):  # Only process .xlsx files
            file_path = os.path.join(directory, filename)
            print(f"\nProcessing: {filename}")
            
            if process_excel_file(file_path):
                success_count += 1
            else:
                failure_count += 1
    
    print(f"\nProcessing complete!")
    print(f"Successfully processed: {success_count} files")
    print(f"Failed to process: {failure_count} files")

if __name__ == "__main__":
    main()